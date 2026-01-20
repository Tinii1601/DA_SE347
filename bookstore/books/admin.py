# books/admin.py
from django.contrib import admin
from django.utils.html import format_html
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Category, Product, Attribute, ProductAttributeValue, ProductImage
from .forms import ExcelImportForm
import openpyxl
import requests
from django.core.files.base import ContentFile

# --- Shared Logic ---

def process_attributes_sheet(wb):
    if "Attributes" in wb.sheetnames:
        ws = wb["Attributes"]
        count = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:
                Attribute.objects.get_or_create(name=row[0])
                count += 1
        return count, "Attributes"
    return 0, None

def process_categories_sheet(wb):
    if "Categories" in wb.sheetnames:
        ws = wb["Categories"]
        count = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[0]:
                name = row[0]
                slug = row[1] if len(row) > 1 and row[1] else None
                description = row[2] if len(row) > 2 and row[2] else ""
                parent_name = row[3] if len(row) > 3 else None
                attr_list_str = row[4] if len(row) > 4 else None
                
                parent = None
                if parent_name:
                    parent = Category.objects.filter(name=parent_name).first()
                
                category, created = Category.objects.update_or_create(
                    name=name,
                    defaults={
                        'slug': slug if slug else "", 
                        'description': description,
                        'parent': parent
                    }
                )

                if attr_list_str:
                    attr_names = [x.strip() for x in str(attr_list_str).split(';') if x.strip()]
                    for attr_name in attr_names:
                        attr_obj, _ = Attribute.objects.get_or_create(name=attr_name)
                        category.attributes.add(attr_obj)

                count += 1
        return count, "Categories"
    return 0, None

def process_products_sheet(wb, request_msg_func=None):
    if "Products" in wb.sheetnames:
        ws = wb["Products"]
        count = 0
        errors = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            try:
                if not row[0] or not row[1]: 
                    continue

                cat_name = row[0]
                name = row[1]
                price = row[2] if len(row) > 2 else 0
                stock = row[3] if len(row) > 3 else 0
                desc = row[4] if len(row) > 4 else ""
                img_url = row[5] if len(row) > 5 else ""
                attrs_str = row[6] if len(row) > 6 else ""
                album_urls = row[7] if len(row) > 7 else ""

                category = Category.objects.filter(name=cat_name).first()
                if not category:
                    errors.append(f"Không tìm thấy danh mục '{cat_name}' cho sản phẩm '{name}'")
                    continue

                product, created = Product.objects.update_or_create(
                    name=name,
                    defaults={
                        'category': category,
                        'price': price,
                        'stock': stock,
                        'description': desc
                    }
                )

                if img_url and str(img_url).strip():
                    try:
                        response = requests.get(img_url, timeout=10)
                        if response.status_code == 200:
                            # Try to guess filename from URL or default to product slug
                            file_name = img_url.split("/")[-1]
                            if not file_name or '.' not in file_name:
                                file_name = f"{product.slug}.jpg"
                            
                            product.cover_image.save(file_name, ContentFile(response.content), save=True)
                    except Exception as e:
                        errors.append(f"Không tải được ảnh cho '{name}': {e}")
                
                if album_urls and str(album_urls).strip():
                    img_urls = str(album_urls).split(';')
                    for idx, url in enumerate(img_urls):
                        url = url.strip()
                        if url:
                            try:
                                res = requests.get(url, timeout=10)
                                if res.status_code == 200:
                                    fname = url.split("/")[-1]
                                    if not fname or '.' not in fname:
                                        fname = f"{product.slug}_album_{idx}.jpg"
                                    
                                    # Create ProductImage
                                    p_img = ProductImage(product=product)
                                    p_img.image.save(fname, ContentFile(res.content), save=True)
                            except Exception as e:
                                errors.append(f"Không tải được ảnh album cho '{name}': {e}")

                if attrs_str:
                    pairs = str(attrs_str).split(';')
                    for pair in pairs:
                        if ':' in pair:
                            attr_name, val = pair.split(':', 1)
                            attr_name = attr_name.strip()
                            val = val.strip()
                            
                            attribute = Attribute.objects.filter(name=attr_name).first()
                            if attribute:
                                ProductAttributeValue.objects.update_or_create(
                                    product=product,
                                    attribute=attribute,
                                    defaults={'value': val}
                                )
                count += 1
            except Exception as e:
                errors.append(f"Lỗi sản phẩm '{row[1]}': {str(e)}")
        
        if request_msg_func and errors:
            for err in errors:
                request_msg_func(err, level="warning")
                
        return count, "Products"
    return 0, None

# --- Admin Classes ---

class ProductAttributeValueInline(admin.TabularInline):
    model = ProductAttributeValue
    extra = 1
    verbose_name = "Giá trị thuộc tính"
    verbose_name_plural = "Các giá trị thuộc tính"

class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1
    verbose_name = "Ảnh album"
    verbose_name_plural = "Album ảnh"

@admin.register(Attribute)
class AttributeAdmin(admin.ModelAdmin):
    list_display = ['name']
    # change_list_template removed to avoid duplication with JS injection

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='books_attribute_import_excel'),
        ]
        return my_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    count, _ = process_attributes_sheet(wb)
                    messages.success(request, f"Đã nhập {count} thuộc tính.")
                except Exception as e:
                    messages.error(request, f"Lỗi: {e}")
                return redirect("admin:books_attribute_changelist")
        else:
            form = ExcelImportForm()
        
        context = {
            "form": form,
            "title": "Nhập Thuộc tính từ Excel",
            "opts": self.model._meta,
            "description": """
                <b>Yêu cầu file Excel:</b><br>
                - Đọc từ sheet tên là: <b>"Attributes"</b><br>
                - Cột A: Tên thuộc tính (Ví dụ: Màu sắc, Kích thước).
            """
        }
        return render(request, "admin/books/import_excel_generic.html", context)


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'slug', 'product_count']
    search_fields = ['name']
    prepopulated_fields = {'slug': ('name',)}
    # change_list_template removed to avoid duplication with JS injection

    def product_count(self, obj):
        return obj.products.count()
    product_count.short_description = "Số sản phẩm"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='books_category_import_excel'),
        ]
        return my_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    count, _ = process_categories_sheet(wb)
                    messages.success(request, f"Đã nhập {count} danh mục.")
                except Exception as e:
                    messages.error(request, f"Lỗi: {e}")
                return redirect("admin:books_category_changelist")
        else:
            form = ExcelImportForm()

        context = {
            "form": form,
            "title": "Nhập Danh mục từ Excel",
            "opts": self.model._meta,
            "description": """
                <b>Yêu cầu file Excel:</b><br>
                - Đọc từ sheet tên là: <b>"Categories"</b><br>
                - Cột A: Tên danh mục (Bắt buộc).<br>
                - Cột B: Slug (Tùy chọn).<br>
                - Cột C: Mô tả (Tùy chọn).<br>
                - Cột D: Tên danh mục cha (Tùy chọn - phải trùng khớp với tên danh mục đã có).<br>
                - Cột E: Danh sách thuộc tính (Tùy chọn - ngăn cách bởi dấu phẩy; ví dụ: <code>Màu sắc; Tác giả</code>).
            """
        }
        return render(request, "admin/books/import_excel_generic.html", context)

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ['name', 'final_price', 'stock', 'is_active', 'cover_image_thumb']
    list_filter = ['is_active', 'category', 'created_at']
    search_fields = ['name']
    prepopulated_fields = {'slug': ('name',)}
    readonly_fields = ['created_at', 'updated_at']
    list_editable = ['is_active']
    inlines = [ProductAttributeValueInline, ProductImageInline]
    # change_list_template removed to avoid duplication with JS injection

    def cover_image_thumb(self, obj):
        if obj.cover_image:
            return format_html(
                '<img src="{}" width="60" height="90" style="object-fit: cover; border-radius: 4px;" />',
                obj.cover_image.url
            )
        return "Chưa có ảnh"
    cover_image_thumb.short_description = "Ảnh SP"

    def final_price(self, obj):
        price = obj.get_final_price()
        return f"{price:,.0f}₫"
    final_price.short_description = "Giá bán"

    def get_urls(self):
        urls = super().get_urls()
        
        # URL for Product-only import and Master import
        my_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='books_product_import_excel'),
            path('import-master/', self.admin_site.admin_view(self.import_master), name='books_master_import_excel'),
        ]
        return my_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    
                    def msg_func(msg, level="error"):
                        if level == "warning": messages.warning(request, msg)
                        else: messages.error(request, msg)

                    count, _ = process_products_sheet(wb, request_msg_func=msg_func)
                    messages.success(request, f"Đã nhập {count} sản phẩm.")
                except Exception as e:
                    messages.error(request, f"Lỗi: {e}")
                return redirect("admin:books_product_changelist")
        else:
            form = ExcelImportForm()

        context = {
            "form": form,
            "title": "Nhập Sản phẩm từ Excel",
            "opts": self.model._meta,
            "description": """
                <b>Yêu cầu file Excel:</b><br>
                - Đọc từ sheet tên là: <b>"Products"</b><br>
                - Cột A: Tên danh mục (Bắt buộc - phải khớp với danh mục đã có).<br>
                - Cột B: Tên sản phẩm (Bắt buộc).<br>
                - Cột C: Giá bán.<br>
                - Cột D: Số lượng tồn kho.<br>
                - Cột E: Mô tả.<br>
                - Cột F: Link Ảnh bìa (URL).<br>
                - Cột G: Các thuộc tính (Định dạng: <code>Màu sắc:Đỏ; Kích thước:XL</code>).<br>
                - Cột H: Link Ảnh Album (URL, ngăn cách bởi dấu chấm phẩy ;).
            """
        }
        return render(request, "admin/books/import_excel_generic.html", context)

    def import_master(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    
                    # Process in order: Attributes -> Categories -> Products
                    c_attrs, _ = process_attributes_sheet(wb)
                    c_cats, _ = process_categories_sheet(wb)
                    
                    def msg_func(msg, level="error"):
                        if level == "warning": messages.warning(request, msg)
                        else: messages.error(request, msg)
                    
                    c_prods, _ = process_products_sheet(wb, request_msg_func=msg_func)
                    
                    messages.success(request, f"Đã nhập tổng hợp: {c_attrs} thuộc tính, {c_cats} danh mục, {c_prods} sản phẩm.")
                except Exception as e:
                    messages.error(request, f"Lỗi: {e}")
                return redirect("admin:app_list", app_label='books')
        else:
            form = ExcelImportForm()

        context = {
            "form": form,
            "title": "Nhập Tổng hợp dữ liệu từ Excel",
            "opts": self.model._meta,
            "description": """
                <b>Yêu cầu file Excel (Chứa 3 Sheet):</b><br><br>
                
                <b>1. Nhập Thuộc tính (Attribute):</b><br>
                - Sheet: <b>"Attributes"</b><br>
                - Cột A: Tên thuộc tính<br><br>
                
                <b>2. Nhập Danh mục (Category):</b><br>
                - Sheet: <b>"Categories"</b><br>
                - Cột A: Tên danh mục | Cột B: Slug | Cột C: Mô tả | Cột D: Danh mục cha | Cột E: Thuộc tính (chấm phẩy ;)<br><br>
                
                <b>3. Nhập Sản phẩm (Product):</b><br>
                - Sheet: <b>"Products"</b><br>
                - Cột A: Tên danh mục | Cột B: Tên SP | Cột C: Giá | Cột D:Số lượng tồn kho | Cột E: Mô tả | Cột F: Link Ảnh | Cột G: Thuộc tính (chấm phẩy ;) | Cột H: Album Ảnh (chấm phẩy ;)<br><br>
            """
        }
        return render(request, "admin/books/import_excel_generic.html", context)


